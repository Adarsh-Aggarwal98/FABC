"""
Import clients from JPATax Pre-tax Consultation Form Excel file
"""
import os
import sys
from datetime import datetime

# Add the app to the path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl #test comment
from app import create_app
from app.extensions import db
from app.modules.user.models import User, Role

def clean_phone(phone):
    """Clean phone number"""
    if phone is None:
        return None
    phone = str(phone).strip()
    # Remove .0 from float conversion
    if phone.endswith('.0'):
        phone = phone[:-2]
    # Add leading 0 if it's a 9-digit Australian number
    if len(phone) == 9 and phone.isdigit():
        phone = '0' + phone
    return phone

def clean_tfn(tfn):
    """Clean TFN - remove spaces, handle float"""
    if tfn is None:
        return None
    tfn = str(tfn).strip()
    if tfn.endswith('.0'):
        tfn = tfn[:-2]
    # Remove spaces
    tfn = tfn.replace(' ', '')
    return tfn

def clean_account_number(acc):
    """Clean account number"""
    if acc is None:
        return None
    acc = str(acc).strip()
    if acc.endswith('.0'):
        acc = acc[:-2]
    return acc

def map_visa_status(visa):
    """Map form visa status to CRM format"""
    if visa is None:
        return None
    visa_lower = visa.lower()
    if 'citizen' in visa_lower:
        return 'citizen'
    elif 'permanent' in visa_lower:
        return 'permanent_resident'
    elif 'student' in visa_lower:
        return 'student'
    elif 'working holiday' in visa_lower or '417' in visa_lower or '462' in visa_lower:
        return 'working_holiday'
    elif 'temporary' in visa_lower:
        return 'temporary_resident'
    else:
        return 'other'

def import_clients(excel_path, dry_run=True):
    """Import clients from Excel file"""
    app = create_app()

    with app.app_context():
        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # Get headers
        headers = {}
        for col in range(1, sheet.max_column + 1):
            headers[sheet.cell(row=1, column=col).value] = col

        # Get or create user role
        user_role = Role.query.filter_by(name=Role.USER).first()
        if not user_role:
            print("ERROR: User role not found. Run flask init-db first.")
            return

        imported = 0
        skipped = 0
        errors = []

        print(f"\n{'='*60}")
        print(f"Importing clients from: {excel_path}")
        print(f"Total rows: {sheet.max_row - 1}")
        print(f"Dry run: {dry_run}")
        print(f"{'='*60}\n")

        for row in range(2, sheet.max_row + 1):
            try:
                # Get email
                email = sheet.cell(row=row, column=headers.get('Email ID')).value
                if not email:
                    # Try Email Address column
                    email = sheet.cell(row=row, column=headers.get('Email Address')).value

                if not email:
                    print(f"Row {row}: Skipped - No email")
                    skipped += 1
                    continue

                email = email.strip().lower()

                # Check if user exists
                existing = User.query.filter_by(email=email).first()
                if existing:
                    print(f"Row {row}: Skipped - {email} already exists")
                    skipped += 1
                    continue

                # Get data from row
                first_name = sheet.cell(row=row, column=headers.get('First Name')).value
                last_name = sheet.cell(row=row, column=headers.get('Last / Family Name')).value
                phone = clean_phone(sheet.cell(row=row, column=headers.get('Mobile Number')).value)
                address = sheet.cell(row=row, column=headers.get('Current Residential Address')).value
                visa_status = map_visa_status(sheet.cell(row=row, column=headers.get('Select your current visa status')).value)
                tfn = clean_tfn(sheet.cell(row=row, column=headers.get('Tax File Number (TFN)')).value)
                dob = sheet.cell(row=row, column=headers.get('Date of Birth')).value
                occupation = sheet.cell(row=row, column=headers.get('Occupation')).value
                bsb = sheet.cell(row=row, column=headers.get('BSB')).value
                account_number = clean_account_number(sheet.cell(row=row, column=headers.get('Account Number')).value)
                account_holder = sheet.cell(row=row, column=headers.get('Account Holder Name')).value
                passport_url = sheet.cell(row=row, column=headers.get('Passport Copy (Both pages)')).value
                bank_statement_url = sheet.cell(row=row, column=headers.get('Bank Statement')).value
                driving_licence_url = sheet.cell(row=row, column=headers.get('Australian Driving Licence (both sides)')).value
                terms_col = [h for h in headers.keys() if h and 'hereby' in h.lower()]
                terms_accepted = False
                if terms_col:
                    terms_value = sheet.cell(row=row, column=headers.get(terms_col[0])).value
                    terms_accepted = terms_value and 'accept' in str(terms_value).lower()

                # Convert date of birth
                if isinstance(dob, datetime):
                    dob = dob.date()
                elif dob:
                    dob = None  # Can't parse string dates reliably

                # Create user
                user = User(
                    email=email,
                    role_id=user_role.id,
                    first_name=first_name.strip() if first_name else None,
                    last_name=last_name.strip() if last_name else None,
                    phone=phone,
                    address=address.strip() if address else None,
                    visa_status=visa_status,
                    tfn=tfn,
                    date_of_birth=dob,
                    occupation=occupation.strip() if occupation else None,
                    bsb=str(bsb).strip() if bsb else None,
                    bank_account_number=account_number,
                    bank_account_holder_name=account_holder.strip() if account_holder else None,
                    passport_url=passport_url,
                    bank_statement_url=bank_statement_url,
                    driving_licence_url=driving_licence_url,
                    terms_accepted=terms_accepted,
                    terms_accepted_at=datetime.utcnow() if terms_accepted else None,
                    is_active=True,
                    is_verified=False,
                    is_first_login=True
                )

                # Set a temporary password (they'll need to reset)
                user.set_password('TempPass123!')

                print(f"Row {row}: {'[DRY RUN] Would import' if dry_run else 'Importing'} {email} ({first_name} {last_name})")

                if not dry_run:
                    db.session.add(user)

                imported += 1

            except Exception as e:
                errors.append(f"Row {row}: {str(e)}")
                print(f"Row {row}: ERROR - {str(e)}")

        if not dry_run:
            db.session.commit()
            print("\nChanges committed to database.")

        print(f"\n{'='*60}")
        print(f"SUMMARY")
        print(f"{'='*60}")
        print(f"Imported: {imported}")
        print(f"Skipped: {skipped}")
        print(f"Errors: {len(errors)}")
        if errors:
            print("\nErrors:")
            for e in errors:
                print(f"  - {e}")
        print(f"{'='*60}\n")

if __name__ == '__main__':
    excel_file = r"D:\Downloads\JPATax_Pre-tax Consultation Form (Responses).xlsx"

    # First do a dry run
    print("Running in DRY RUN mode first...")
    import_clients(excel_file, dry_run=True)

    # Ask for confirmation
    response = input("\nDo you want to import for real? (yes/no): ")
    if response.lower() == 'yes':
        import_clients(excel_file, dry_run=False)
    else:
        print("Import cancelled.")
