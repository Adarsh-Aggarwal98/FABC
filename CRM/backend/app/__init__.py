import os
import logging
import sys
from flask import Flask, jsonify
from app.config import config
from app.extensions import db, migrate, jwt, cors
from app.modules.metrics import init_metrics, track_login, track_token_operation


def setup_logging(app):
    """Configure logging to output to stdout for docker logs"""
    # Remove default handlers
    app.logger.handlers = []

    # Create stdout handler
    handler = logging.StreamHandler(sys.stdout)
    handler.setLevel(logging.DEBUG if app.debug else logging.INFO)

    # Create formatter
    formatter = logging.Formatter(
        '[%(asctime)s] %(levelname)s in %(module)s: %(message)s'
    )
    handler.setFormatter(formatter)

    # Add handler to app logger
    app.logger.addHandler(handler)
    app.logger.setLevel(logging.DEBUG if app.debug else logging.INFO)

    # Also configure root logger for libraries
    logging.getLogger().addHandler(handler)
    logging.getLogger().setLevel(logging.INFO)


def create_app(config_name=None):
    """Application factory for creating Flask app"""

    if config_name is None:
        config_name = os.getenv('FLASK_ENV', 'development')

    app = Flask(__name__)
    app.config.from_object(config[config_name])
    app.url_map.strict_slashes = False  # Fix 308 redirects

    # Setup logging for docker/gunicorn
    setup_logging(app)

    # Initialize extensions
    db.init_app(app)
    migrate.init_app(app, db)
    jwt.init_app(app)

    # =========================================================================
    # DATABASE SETUP - Create tables from SQLAlchemy models (if they don't exist)
    # This preserves existing data while ensuring schema matches models
    # =========================================================================
    with app.app_context():
        # Import all models so SQLAlchemy knows about them
        from app.modules.user.models import Role, User
        from app.modules.company.models import Company
        from app.modules.services.models import Service, ServiceRequest, Query, Invoice, Task, ClientServicePricing
        from app.modules.notifications.models import Notification, EmailTemplate
        from app.modules.forms.models import Form, FormQuestion, FormResponse
        from app.modules.documents.models import Document
        from app.modules.tags.models import ClientTag
        from app.modules.audit.models import ActivityLog
        from app.modules.client_entity.models import ClientEntity
        from app.modules.leads.models import Lead
        from app.modules.letters.models import AuditLetter
        from app.modules.smsf_data_sheet.models import SMSFDataSheet
        from app.modules.blog.models import Blog
        from app.modules.ato_alerts.models import AtoAlert
        import bcrypt

        # Create tables that don't exist (preserves existing data)
        # Wrapped to handle race condition when multiple gunicorn workers start simultaneously
        try:
            db.create_all()
            app.logger.info('Database tables verified/created from SQLAlchemy models.')
        except Exception as e:
            db.session.rollback()
            if 'already exists' in str(e) or 'duplicate key' in str(e):
                app.logger.info('Database tables already exist (another worker already initialised them).')
            else:
                raise

        # Seed essential data only if missing
        try:
            # Create roles if missing
            Role.get_or_create_default_roles()

            # Create company if missing
            company = Company.query.filter_by(name='AusSuperSource').first()
            if not company:
                company = Company(
                    name='AusSuperSource',
                    trading_name='AusSuperSource Pty Ltd',
                    email='info@aussupersource.com.au',
                    phone='',
                    address_line1='Sydney',
                    city='Sydney',
                    state='NSW',
                    postcode='2000',
                    country='Australia',
                    plan_type='enterprise',
                    max_users=100,
                    max_clients=10000,
                    # AusSuperSource branding
                    primary_color='#1B72BE',
                    secondary_color='#1A2E5A',
                    tertiary_color='#3a9ef0',
                    logo_url='/assets/aussupersource-logo.png'
                )
                db.session.add(company)
                db.session.flush()
                app.logger.info(f'Company created: {company.name}')

            # Create admin user if missing
            admin_user = User.query.filter_by(email='aggarwal.adarsh98@gmail.com').first()
            if not admin_user:
                super_admin_role = Role.query.filter_by(name=Role.SUPER_ADMIN).first()
                admin_user = User(
                    email='aggarwal.adarsh98@gmail.com',
                    role_id=super_admin_role.id,
                    company_id=company.id,
                    first_name='Adarsh',
                    last_name='Aggarwal',
                    is_active=True,
                    is_verified=True,
                    is_first_login=False,
                    two_fa_enabled=False
                )
                # Set password using bcrypt (Big@200650078296)
                admin_user.password_hash = bcrypt.hashpw('Big@200650078296'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                db.session.add(admin_user)
                app.logger.info('Admin user created: aggarwal.adarsh98@gmail.com')

            db.session.commit()

            # Seed default blogs if empty
            try:
                if Blog.query.count() == 0:
                    default_blogs = [
                        Blog(
                            title='2025 SMSF Compliance Changes: What Accountants Need to Know',
                            slug='2025-smsf-compliance-changes',
                            excerpt='Stay ahead of the latest regulatory updates affecting SMSF audits and compliance.',
                            category='Compliance',
                            author='Yateender Gupta',
                            read_time='5 min read',
                            image='https://images.unsplash.com/photo-1554224155-8d04cb21cd6c?w=600&h=400&fit=crop',
                            featured=True,
                        ),
                        Blog(
                            title='Partner Auditing: How to Scale Your SMSF Practice',
                            slug='partner-auditing-scale-practice',
                            excerpt='Discover how partnering with a specialist SMSF audit provider can help you grow.',
                            category='Practice Growth',
                            author='Sharat Sharma',
                            read_time='4 min read',
                            image='https://images.unsplash.com/photo-1460925895917-afdab827c52f?w=600&h=400&fit=crop',
                        ),
                    ]
                    for b in default_blogs:
                        db.session.add(b)
                    db.session.commit()
                    app.logger.info('Default blogs seeded.')
            except Exception as e:
                db.session.rollback()
                app.logger.warning(f'Could not seed blogs: {e}')

            # Seed default ATO alerts if empty
            try:
                if AtoAlert.query.count() == 0:
                    default_alerts = [
                        AtoAlert(
                            title='SMSF annual return deadline extended to 28 February 2026',
                            type='update',
                            link='https://www.ato.gov.au/super/self-managed-super-funds/',
                            priority=1,
                        ),
                        AtoAlert(
                            title='New contribution caps for 2025-26: Concessional $30,000',
                            type='alert',
                            link='https://www.ato.gov.au/super/self-managed-super-funds/contributions-and-rollovers/',
                            priority=2,
                        ),
                    ]
                    for a in default_alerts:
                        db.session.add(a)
                    db.session.commit()
                    app.logger.info('Default ATO alerts seeded.')
            except Exception as e:
                db.session.rollback()
                app.logger.warning(f'Could not seed ATO alerts: {e}')

        except Exception as e:
            db.session.rollback()
            app.logger.warning(f'Could not seed initial data: {e}')
    # =========================================================================
    # Get CORS origins from config (supports environment-based configuration)
    cors_origins = app.config.get('CORS_ORIGINS', [
        "http://localhost:5173",
        "http://localhost:3000",
        "http://localhost:3001",
    ])
    # Add production origins if not already present
    production_origins = [
        "https://aussupersource.com.au",
        "https://www.aussupersource.com.au",
        "https://crm.aussupersource.com.au",
        "https://api.aussupersource.com.au"
    ]
    for origin in production_origins:
        if origin not in cors_origins:
            cors_origins.append(origin)

    cors.init_app(app, resources={
        r"/api/*": {
            "origins": cors_origins,
            "methods": ["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
            "allow_headers": ["Content-Type", "Authorization"],
            "supports_credentials": True
        }
    })

    # Initialize Prometheus metrics
    try:
        init_metrics(app)
        app.logger.info('Prometheus metrics initialized successfully')
    except Exception as e:
        app.logger.warning(f'Could not initialize Prometheus metrics: {e}')

    # Initialize background job scheduler (APScheduler)
    # Only start in main process, not in reloader or CLI commands
    if not app.debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        try:
            from app.jobs import init_scheduler
            init_scheduler(app)
        except ImportError as e:
            app.logger.warning(f'APScheduler not available: {e}. Install with: pip install apscheduler')
        except Exception as e:
            app.logger.error(f'Failed to start scheduler: {e}')

    # JWT error handlers
    @jwt.expired_token_loader
    def expired_token_callback(jwt_header, jwt_payload):
        return jsonify({
            'success': False,
            'error': 'Token has expired',
            'message': 'Please login again'
        }), 401

    @jwt.invalid_token_loader
    def invalid_token_callback(error):
        return jsonify({
            'success': False,
            'error': 'Invalid token',
            'message': str(error)
        }), 401

    @jwt.unauthorized_loader
    def missing_token_callback(error):
        return jsonify({
            'success': False,
            'error': 'Authorization required',
            'message': 'Missing access token'
        }), 401

    # Register blueprints
    from app.modules.user import auth_bp, user_bp
    from app.modules.services import services_bp, requests_bp, workflow_bp, renewals_bp, status_bp, invoices_bp, task_bp, client_pricing_bp
    from app.modules.notifications import notifications_bp
    from app.modules.forms import forms_bp
    from app.modules.documents import documents_bp
    from app.modules.company import company_bp
    from app.modules.tags import tags_bp
    from app.modules.audit import audit_bp
    from app.modules.search import search_bp
    from app.modules.analytics import analytics_bp
    from app.modules.payments import payments_bp
    from app.modules.imports.routes import import_bp
    from app.modules.client_entity import client_entity_bp
    from app.modules.leads import leads_bp

    app.register_blueprint(auth_bp, url_prefix='/api/auth')
    app.register_blueprint(user_bp, url_prefix='/api/users')
    app.register_blueprint(services_bp, url_prefix='/api/services')
    app.register_blueprint(requests_bp, url_prefix='/api/requests')
    app.register_blueprint(workflow_bp, url_prefix='/api/workflows')
    app.register_blueprint(renewals_bp, url_prefix='/api/renewals')
    app.register_blueprint(status_bp)  # Already has /api/statuses prefix
    app.register_blueprint(task_bp)  # Already has /api/tasks prefix
    app.register_blueprint(invoices_bp, url_prefix='/api/services/invoices')
    app.register_blueprint(client_pricing_bp)  # Already has /api/client-pricing prefix
    app.register_blueprint(notifications_bp, url_prefix='/api/notifications')
    app.register_blueprint(forms_bp, url_prefix='/api/forms')
    app.register_blueprint(documents_bp, url_prefix='/api/documents')
    app.register_blueprint(company_bp, url_prefix='/api/companies')
    app.register_blueprint(tags_bp, url_prefix='/api/tags')
    app.register_blueprint(audit_bp, url_prefix='/api/activity')
    app.register_blueprint(search_bp, url_prefix='/api/search')
    app.register_blueprint(analytics_bp, url_prefix='/api/analytics')
    app.register_blueprint(payments_bp, url_prefix='/api/payments')
    app.register_blueprint(import_bp)  # Already has /api/imports prefix
    app.register_blueprint(client_entity_bp, url_prefix='/api/client-entities')

    # Register Leads module (website form submissions)
    app.register_blueprint(leads_bp)  # Already has /api/leads prefix

    # Register Letters module (SMSF Engagement & Representation letter generation)
    from app.modules.letters import letters_bp
    app.register_blueprint(letters_bp)  # Already has /api/letters prefix

    # Register SMSF Data Sheet module
    from app.modules.smsf_data_sheet import data_sheet_bp
    app.register_blueprint(data_sheet_bp)  # Already has /api/smsf-data-sheets prefix

    # Register Blog module (public articles for website)
    from app.modules.blog import blog_bp
    app.register_blueprint(blog_bp)

    # Register ATO Alerts module (public regulatory alerts for website)
    from app.modules.ato_alerts import ato_alerts_bp
    app.register_blueprint(ato_alerts_bp)

    # Register Client Portal module (invite + dashboard for SMSF clients)
    from app.modules.client_portal import client_portal_bp
    app.register_blueprint(client_portal_bp)  # Already has /api/client-portal prefix

    # Register Google Drive integration blueprint
    from app.modules.integrations.google_drive.routes import google_drive_bp
    app.register_blueprint(google_drive_bp)

    # Error handlers
    @app.errorhandler(404)
    def not_found(error):
        return jsonify({
            'success': False,
            'error': 'Resource not found'
        }), 404

    @app.errorhandler(500)
    def internal_error(error):
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500

    # Health check endpoint
    @app.route('/api/health')
    def health_check():
        return jsonify({
            'status': 'healthy',
            'message': 'Accountant CRM API is running'
        })

    # CLI commands
    @app.cli.command('init-db')
    def init_db():
        """Initialize database with default data"""
        from app.modules.user.models import Role, User
        from app.modules.company.models import Company

        # Create default roles
        Role.get_or_create_default_roles()
        print('Default roles created.')

        # Get roles
        super_admin_role = Role.query.filter_by(name=Role.SUPER_ADMIN).first()
        admin_role = Role.query.filter_by(name=Role.ADMIN).first()
        accountant_role = Role.query.filter_by(name=Role.ACCOUNTANT).first()
        user_role = Role.query.filter_by(name=Role.USER).first()

        # Create super admin if not exists
        existing_super_admin = User.query.filter_by(email='admin@example.com').first()
        if not existing_super_admin:
            super_admin = User(
                email='admin@example.com',
                role_id=super_admin_role.id,
                first_name='Super',
                last_name='Admin',
                is_first_login=False,
                is_verified=True,
                two_fa_enabled=False  # Disable 2FA for easy testing
            )
            super_admin.set_password('Admin@123')
            db.session.add(super_admin)
            db.session.flush()
            print('Super Admin created:')
            print('  Email: admin@example.com')
            print('  Password: Admin@123')
        else:
            super_admin = existing_super_admin
            print('Super admin already exists.')

        # Create sample company if not exists
        existing_company = Company.query.filter_by(name='Demo Accounting Practice').first()
        if not existing_company:
            company = Company(
                name='Demo Accounting Practice',
                trading_name='Demo Tax Services',
                abn='12 345 678 901',
                email='info@demopractice.com',
                phone='+61 2 1234 5678',
                address_line1='123 Main Street',
                city='Sydney',
                state='NSW',
                postcode='2000',
                country='Australia',
                plan_type='standard',
                max_users=10,
                max_clients=100
            )
            db.session.add(company)
            db.session.flush()
            print('Sample company created: Demo Accounting Practice')
        else:
            company = existing_company
            print('Sample company already exists.')

        # Create practice admin (company owner) if not exists
        existing_practice_admin = User.query.filter_by(email='practiceadmin@example.com').first()
        if not existing_practice_admin:
            practice_admin = User(
                email='practiceadmin@example.com',
                role_id=admin_role.id,
                company_id=company.id,
                first_name='Practice',
                last_name='Admin',
                is_first_login=False,
                is_verified=True,
                two_fa_enabled=False
            )
            practice_admin.set_password('Practice@123')
            db.session.add(practice_admin)
            db.session.flush()

            # Set as company owner
            company.owner_id = practice_admin.id

            print('Practice Admin created:')
            print('  Email: practiceadmin@example.com')
            print('  Password: Practice@123')
        else:
            print('Practice admin already exists.')

        # Create accountant if not exists
        existing_accountant = User.query.filter_by(email='accountant@example.com').first()
        if not existing_accountant:
            accountant = User(
                email='accountant@example.com',
                role_id=accountant_role.id,
                company_id=company.id,
                first_name='John',
                last_name='Accountant',
                is_first_login=False,
                is_verified=True,
                two_fa_enabled=False
            )
            accountant.set_password('Accountant@123')
            db.session.add(accountant)
            db.session.flush()
            print('Accountant created:')
            print('  Email: accountant@example.com')
            print('  Password: Accountant@123')
        else:
            accountant = existing_accountant
            print('Accountant already exists.')

        # Create test client user if not exists
        existing_client = User.query.filter_by(email='client@example.com').first()
        if not existing_client:
            client = User(
                email='client@example.com',
                role_id=user_role.id,
                company_id=company.id,
                first_name='Test',
                last_name='Client',
                phone='+61 412 345 678',
                address='456 Test Street, Sydney NSW 2000',
                is_first_login=False,
                is_verified=True,
                two_fa_enabled=False
            )
            client.set_password('Client@123')
            db.session.add(client)
            print('Test Client created:')
            print('  Email: client@example.com')
            print('  Password: Client@123')
        else:
            print('Test client already exists.')

        db.session.commit()
        print('\n--- Default Users Summary ---')
        print('Super Admin:     admin@example.com / Admin@123 (no company)')
        print('Practice Admin:  practiceadmin@example.com / Practice@123')
        print('Accountant:      accountant@example.com / Accountant@123')
        print('Test Client:     client@example.com / Client@123')
        print('\nSample Company: Demo Accounting Practice')
        print('  - Practice Admin is the owner')
        print('  - Accountant and Client are members')

    @app.cli.command('create-sample-services')
    def create_sample_services():
        """Create services with associated forms"""
        from app.modules.services.models import Service
        from app.modules.forms.models import Form, FormQuestion
        from app.modules.user.models import User, Role

        # Get super admin for form creator
        super_admin = User.query.join(Role).filter(Role.name == Role.SUPER_ADMIN).first()
        creator_id = super_admin.id if super_admin else None

        # Define services with their forms
        services_with_forms = [
            {
                'name': 'Individual Tax Return',
                'description': 'Annual individual tax return preparation and lodgement with ATO',
                'category': 'Tax',
                'base_price': 350.00,
                'form': {
                    'name': 'Individual Tax Return Information',
                    'description': 'Please provide the following information for your tax return',
                    'questions': [
                        {'question_text': 'What is your Tax File Number (TFN)?', 'question_type': 'text', 'is_required': True, 'placeholder': 'XXX XXX XXX'},
                        {'question_text': 'What is your date of birth?', 'question_type': 'date', 'is_required': True},
                        {'question_text': 'What is your occupation?', 'question_type': 'text', 'is_required': True},
                        {'question_text': 'Did you have any income from employment?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes', 'No']},
                        {'question_text': 'Upload your PAYG Payment Summary / Income Statement', 'question_type': 'file', 'is_required': False, 'help_text': 'You can find this in your myGov account'},
                        {'question_text': 'Did you have any bank interest income?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes', 'No']},
                        {'question_text': 'Did you have any dividend income?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes', 'No']},
                        {'question_text': 'Do you have private health insurance?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes', 'No']},
                        {'question_text': 'Upload your Private Health Insurance Statement', 'question_type': 'file', 'is_required': False},
                        {'question_text': 'What work-related expenses do you want to claim?', 'question_type': 'multiselect', 'is_required': False, 'options': ['Work from home expenses', 'Vehicle/Travel expenses', 'Uniform/Clothing', 'Tools and equipment', 'Self-education', 'Other']},
                        {'question_text': 'Please provide details of any other deductions', 'question_type': 'textarea', 'is_required': False, 'placeholder': 'List any additional deductions you want to claim'},
                    ]
                }
            },
            {
                'name': 'Business Activity Statement (BAS)',
                'description': 'Quarterly or monthly BAS preparation and lodgement',
                'category': 'Tax',
                'base_price': 150.00,
                'form': {
                    'name': 'BAS Information Form',
                    'description': 'Provide your business activity details for the period',
                    'questions': [
                        {'question_text': 'What is your ABN?', 'question_type': 'text', 'is_required': True, 'placeholder': 'XX XXX XXX XXX'},
                        {'question_text': 'What period is this BAS for?', 'question_type': 'select', 'is_required': True, 'options': ['July-September (Q1)', 'October-December (Q2)', 'January-March (Q3)', 'April-June (Q4)', 'Monthly']},
                        {'question_text': 'Are you registered for GST?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes', 'No']},
                        {'question_text': 'Total sales/income for the period (including GST)', 'question_type': 'number', 'is_required': True, 'placeholder': '0.00'},
                        {'question_text': 'Total purchases/expenses for the period (including GST)', 'question_type': 'number', 'is_required': True, 'placeholder': '0.00'},
                        {'question_text': 'Do you have PAYG withholding obligations?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes', 'No']},
                        {'question_text': 'Total wages paid for the period', 'question_type': 'number', 'is_required': False, 'placeholder': '0.00'},
                        {'question_text': 'Total PAYG withheld from wages', 'question_type': 'number', 'is_required': False, 'placeholder': '0.00'},
                        {'question_text': 'Upload your sales report/invoices', 'question_type': 'file', 'is_required': False},
                        {'question_text': 'Upload your purchase invoices/receipts', 'question_type': 'file', 'is_required': False},
                        {'question_text': 'Any additional notes or information', 'question_type': 'textarea', 'is_required': False},
                    ]
                }
            },
            {
                'name': 'Investment Rental Property',
                'description': 'Tax return schedule for rental property income and deductions',
                'category': 'Tax',
                'base_price': 200.00,
                'form': {
                    'name': 'Rental Property Information',
                    'description': 'Details about your investment property for tax purposes',
                    'questions': [
                        {'question_text': 'Property address', 'question_type': 'textarea', 'is_required': True, 'placeholder': 'Full address of the rental property'},
                        {'question_text': 'When did you purchase this property?', 'question_type': 'date', 'is_required': True},
                        {'question_text': 'What was the purchase price?', 'question_type': 'number', 'is_required': True, 'placeholder': '0.00'},
                        {'question_text': 'How many weeks was the property rented this year?', 'question_type': 'number', 'is_required': True, 'placeholder': '52'},
                        {'question_text': 'Total rental income received', 'question_type': 'number', 'is_required': True, 'placeholder': '0.00'},
                        {'question_text': 'Is the property managed by an agent?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes - Full management', 'Yes - Letting only', 'No - Self managed']},
                        {'question_text': 'Property management fees paid', 'question_type': 'number', 'is_required': False, 'placeholder': '0.00'},
                        {'question_text': 'Council rates paid', 'question_type': 'number', 'is_required': False, 'placeholder': '0.00'},
                        {'question_text': 'Water rates paid', 'question_type': 'number', 'is_required': False, 'placeholder': '0.00'},
                        {'question_text': 'Insurance premium paid', 'question_type': 'number', 'is_required': False, 'placeholder': '0.00'},
                        {'question_text': 'Interest on loan paid', 'question_type': 'number', 'is_required': False, 'placeholder': '0.00'},
                        {'question_text': 'Repairs and maintenance costs', 'question_type': 'number', 'is_required': False, 'placeholder': '0.00'},
                        {'question_text': 'Upload rental income statement from agent', 'question_type': 'file', 'is_required': False},
                        {'question_text': 'Upload depreciation schedule (if available)', 'question_type': 'file', 'is_required': False},
                        {'question_text': 'Any capital improvements made this year?', 'question_type': 'textarea', 'is_required': False, 'help_text': 'List any renovations or improvements with costs'},
                    ]
                }
            },
            {
                'name': 'Company Incorporation',
                'description': 'New company registration with ASIC and initial setup',
                'category': 'Registration',
                'base_price': 800.00,
                'form': {
                    'name': 'Company Registration Form',
                    'description': 'Information required to register your new company',
                    'questions': [
                        {'question_text': 'Proposed company name (Option 1)', 'question_type': 'text', 'is_required': True, 'help_text': 'Must end with Pty Ltd'},
                        {'question_text': 'Proposed company name (Option 2)', 'question_type': 'text', 'is_required': False, 'help_text': 'Backup name in case first choice is taken'},
                        {'question_text': 'Company type', 'question_type': 'radio', 'is_required': True, 'options': ['Proprietary Limited (Pty Ltd)', 'Public Company (Ltd)']},
                        {'question_text': 'Principal business activity', 'question_type': 'text', 'is_required': True, 'placeholder': 'e.g., IT Consulting, Construction, Retail'},
                        {'question_text': 'Registered office address', 'question_type': 'textarea', 'is_required': True, 'help_text': 'Must be a physical address in Australia'},
                        {'question_text': 'Principal place of business', 'question_type': 'textarea', 'is_required': True},
                        {'question_text': 'Number of directors', 'question_type': 'select', 'is_required': True, 'options': ['1', '2', '3', '4', '5+']},
                        {'question_text': 'Director 1 - Full name', 'question_type': 'text', 'is_required': True},
                        {'question_text': 'Director 1 - Date of birth', 'question_type': 'date', 'is_required': True},
                        {'question_text': 'Director 1 - Residential address', 'question_type': 'textarea', 'is_required': True},
                        {'question_text': 'Director 1 - Place of birth (City, Country)', 'question_type': 'text', 'is_required': True},
                        {'question_text': 'Director 2 - Full name (if applicable)', 'question_type': 'text', 'is_required': False},
                        {'question_text': 'Number of shares to be issued', 'question_type': 'number', 'is_required': True, 'placeholder': '100'},
                        {'question_text': 'Share price', 'question_type': 'number', 'is_required': True, 'placeholder': '1.00'},
                        {'question_text': 'Shareholder 1 - Name', 'question_type': 'text', 'is_required': True},
                        {'question_text': 'Shareholder 1 - Number of shares', 'question_type': 'number', 'is_required': True},
                        {'question_text': 'Do you need GST registration?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes', 'No', 'Not sure']},
                        {'question_text': 'Do you need to register as an employer?', 'question_type': 'radio', 'is_required': True, 'options': ['Yes', 'No', 'Not sure']},
                        {'question_text': 'Upload ID for all directors', 'question_type': 'file', 'is_required': True},
                    ]
                }
            }
        ]

        for svc_data in services_with_forms:
            existing_service = Service.query.filter_by(name=svc_data['name']).first()
            if existing_service:
                print(f"Service '{svc_data['name']}' already exists, skipping...")
                continue

            # Create the form first (as default template)
            form_data = svc_data.pop('form')
            form = Form(
                name=form_data['name'],
                description=form_data['description'],
                form_type='service',
                created_by_id=creator_id,
                is_default=True  # Mark as default template for cloning
            )
            db.session.add(form)
            db.session.flush()  # Get form ID

            # Create questions
            for idx, q_data in enumerate(form_data['questions']):
                question = FormQuestion(
                    form_id=form.id,
                    question_text=q_data['question_text'],
                    question_type=q_data['question_type'],
                    is_required=q_data.get('is_required', False),
                    options=q_data.get('options'),
                    placeholder=q_data.get('placeholder'),
                    help_text=q_data.get('help_text'),
                    order=idx
                )
                db.session.add(question)

            # Create the service with form linked (marked as default/system service)
            service = Service(
                name=svc_data['name'],
                description=svc_data['description'],
                category=svc_data['category'],
                base_price=svc_data.get('base_price'),
                form_id=form.id,
                is_default=True  # Mark as default service
            )
            db.session.add(service)
            print(f"Created default service: {svc_data['name']} with form: {form_data['name']}")

        db.session.commit()
        print('\nAll services with forms created successfully!')

        # Seed comprehensive service forms and auto-link to services
        # This is idempotent - safe to run multiple times
        from app.modules.forms.seed_forms import seed_all_service_forms
        seed_all_service_forms(creator_id)

        # Mark all existing services as default (for upgrade scenarios)
        all_services = Service.query.all()
        for svc in all_services:
            if not svc.is_default:
                svc.is_default = True
        db.session.commit()
        print(f'Marked {len(all_services)} services as default services')

    @app.cli.command('sync-forms')
    def sync_forms():
        """Sync form definitions and form-service links (idempotent).

        Safe to run anytime - creates missing forms, updates existing ones
        (if no responses), and ensures all form-service links are correct.
        """
        from app.modules.forms.seed_forms import seed_all_service_forms
        from app.modules.user.models import User, Role

        super_admin = User.query.join(Role).filter(Role.name == Role.SUPER_ADMIN).first()
        creator_id = super_admin.id if super_admin else None
        seed_all_service_forms(creator_id)

    @app.cli.command('seed-email-templates')
    def seed_email_templates():
        """Seed default email templates"""
        from app.modules.notifications.models import EmailTemplate

        default_templates = [
            {
                'name': 'Welcome Email',
                'slug': 'welcome',
                'subject': 'Welcome to {company_name}!',
                'body_html': '''<p>Dear {client_name},</p>

<p>Welcome to {company_name}! We're excited to have you as our client.</p>

<p>Your account has been created and you can now access our client portal to:</p>
<ul>
    <li>Request services</li>
    <li>Upload documents securely</li>
    <li>Track the progress of your requests</li>
    <li>Communicate with your accountant</li>
</ul>

<p><strong>Login Details:</strong><br>
Email: {client_email}<br>
Portal: {portal_link}</p>

<p>If you have any questions, please don't hesitate to reach out.</p>

<p>Best regards,<br>
{company_name}</p>''',
                'variables': ['client_name', 'client_email', 'company_name', 'portal_link']
            },
            {
                'name': 'Invoice Email',
                'slug': 'invoice',
                'subject': 'Invoice #{invoice_number} from {company_name}',
                'body_html': '''<p>Dear {client_name},</p>

<p>Please find attached your invoice for {service_name}.</p>

<p><strong>Invoice Details:</strong><br>
Invoice Number: {invoice_number}<br>
Amount Due: {amount}<br>
Due Date: {due_date}</p>

<p><strong>Payment Terms:</strong><br>
{payment_terms}</p>

<p><strong>Bank Details:</strong><br>
{bank_details}</p>

<p>You can also pay online using this link: <a href="{payment_link}">Pay Now</a></p>

<p>{invoice_notes}</p>

<p>{invoice_footer}</p>

<p>Best regards,<br>
{company_name}</p>''',
                'variables': ['client_name', 'service_name', 'invoice_number', 'amount', 'due_date', 'payment_terms', 'bank_details', 'payment_link', 'invoice_notes', 'invoice_footer', 'company_name']
            },
            {
                'name': 'Payment Reminder',
                'slug': 'payment_reminder',
                'subject': 'Payment Reminder - Invoice #{invoice_number}',
                'body_html': '''<p>Dear {client_name},</p>

<p>This is a friendly reminder that payment for Invoice #{invoice_number} is due.</p>

<p><strong>Invoice Details:</strong><br>
Service: {service_name}<br>
Amount Due: {amount}<br>
Due Date: {due_date}</p>

<p>If you have already made the payment, please disregard this email.</p>

<p>To make a payment, please use one of the following methods:</p>

<p><strong>Online Payment:</strong><br>
<a href="{payment_link}">Click here to pay securely online</a></p>

<p><strong>Bank Transfer:</strong><br>
{bank_details}</p>

<p>Please include your invoice number as the payment reference.</p>

<p>If you have any questions about this invoice, please contact us.</p>

<p>Best regards,<br>
{company_name}</p>''',
                'variables': ['client_name', 'invoice_number', 'service_name', 'amount', 'due_date', 'payment_link', 'bank_details', 'company_name']
            },
            {
                'name': 'Document Request',
                'slug': 'document_request',
                'subject': 'Documents Required for {service_name}',
                'body_html': '''<p>Dear {client_name},</p>

<p>To proceed with your {service_name}, we require the following documents:</p>

<p>{document_list}</p>

<p>Please upload these documents through our secure portal: <a href="{portal_link}">Upload Documents</a></p>

<p>If you have any questions about what's required, please don't hesitate to ask.</p>

<p>Best regards,<br>
{accountant_name}<br>
{company_name}</p>''',
                'variables': ['client_name', 'service_name', 'document_list', 'portal_link', 'accountant_name', 'company_name']
            },
            {
                'name': 'Service Completed',
                'slug': 'service_completed',
                'subject': 'Your {service_name} is Complete',
                'body_html': '''<p>Dear {client_name},</p>

<p>Great news! Your {service_name} has been completed.</p>

<p>{completion_notes}</p>

<p>You can view the details and download any relevant documents from your client portal: <a href="{portal_link}">View Details</a></p>

<p>If you have any questions, please don't hesitate to reach out.</p>

<p>Thank you for choosing {company_name}.</p>

<p>Best regards,<br>
{accountant_name}<br>
{company_name}</p>''',
                'variables': ['client_name', 'service_name', 'completion_notes', 'portal_link', 'accountant_name', 'company_name']
            },
            {
                'name': 'Query Response Required',
                'slug': 'query_raised',
                'subject': 'Response Required - {service_name}',
                'body_html': '''<p>Dear {client_name},</p>

<p>We have a question regarding your {service_name} and need your response to proceed.</p>

<p><strong>Query:</strong><br>
{query_message}</p>

<p>Please log in to your portal to respond: <a href="{portal_link}">Respond Now</a></p>

<p>Best regards,<br>
{accountant_name}<br>
{company_name}</p>''',
                'variables': ['client_name', 'service_name', 'query_message', 'portal_link', 'accountant_name', 'company_name']
            },
            {
                'name': 'Annual Tax Reminder',
                'slug': 'annual_reminder',
                'subject': 'It\'s Time to Lodge Your Tax Return - {tax_year}',
                'body_html': '''<p>Dear {client_name},</p>

<p>It's that time of year again! The {tax_year} financial year has ended and it's time to prepare your tax return.</p>

<p><strong>Key Dates:</strong><br>
- Tax returns for individuals are due by 31 October {due_year}<br>
- If you're registered with a tax agent (that's us!), you may have an extended deadline</p>

<p><strong>What you need to gather:</strong></p>
<ul>
    <li>Income statements from all employers</li>
    <li>Bank interest statements</li>
    <li>Dividend statements</li>
    <li>Private health insurance statement</li>
    <li>Receipts for work-related expenses</li>
    <li>Investment property documents (if applicable)</li>
</ul>

<p>Ready to get started? <a href="{portal_link}">Request your tax return now</a></p>

<p>Best regards,<br>
{company_name}</p>''',
                'variables': ['client_name', 'tax_year', 'due_year', 'portal_link', 'company_name']
            },
            {
                'name': 'Password Reset',
                'slug': 'password_reset',
                'subject': 'Password Reset Request - {company_name}',
                'body_html': '''<p>Dear {client_name},</p>

<p>We received a request to reset your password for your {company_name} account.</p>

<p>Click the link below to reset your password:</p>

<p><a href="{reset_link}">Reset My Password</a></p>

<p>This link will expire in 24 hours.</p>

<p>If you didn't request a password reset, please ignore this email or contact us if you have concerns.</p>

<p>Best regards,<br>
{company_name}</p>''',
                'variables': ['client_name', 'reset_link', 'company_name']
            }
        ]

        for template_data in default_templates:
            existing = EmailTemplate.query.filter_by(slug=template_data['slug'], company_id=None).first()
            if existing:
                print(f"Template '{template_data['name']}' already exists, skipping...")
                continue

            template = EmailTemplate(
                name=template_data['name'],
                slug=template_data['slug'],
                subject=template_data['subject'],
                body_html=template_data['body_html'],
                variables=template_data['variables'],
                company_id=None,  # System template
                is_active=True
            )
            db.session.add(template)
            print(f"Created template: {template_data['name']}")

        db.session.commit()
        print('\nDefault email templates seeded successfully!')

    @app.cli.command('create-demo-data')
    def create_demo_data():
        """Create demo user and sample requests for client demo"""
        from app.modules.user.models import User, Role
        from app.modules.company.models import Company
        from app.modules.services.models import Service, ServiceRequest, Query
        from app.modules.forms.models import Form, FormResponse
        from datetime import datetime, timedelta
        import random

        # Get roles
        user_role = Role.query.filter_by(name=Role.USER).first()
        accountant_role = Role.query.filter_by(name=Role.ACCOUNTANT).first()

        if not user_role:
            print("Error: Roles not found. Run 'flask init-db' first.")
            return

        # Get the sample company
        company = Company.query.filter_by(name='Demo Accounting Practice').first()
        if not company:
            print("Error: Sample company not found. Run 'flask init-db' first.")
            return

        # Create demo user (client)
        demo_user = User.query.filter_by(email='demo@example.com').first()
        if not demo_user:
            demo_user = User(
                email='demo@example.com',
                role_id=user_role.id,
                company_id=company.id,
                first_name='James',
                last_name='Wilson',
                phone='0412 345 678',
                address='42 Demo Street, Sydney NSW 2000',
                is_first_login=False,
                is_verified=True,
                two_fa_enabled=False
            )
            demo_user.set_password('Demo@123')
            db.session.add(demo_user)
            db.session.flush()
            print('Demo user (client) created:')
            print('  Email: demo@example.com')
            print('  Password: Demo@123')
        else:
            print('Demo user already exists.')

        # Get accountant for assignments
        accountant = User.query.filter_by(email='accountant@example.com').first()

        # Get services
        services = Service.query.all()
        if not services:
            print('No services found. Run "flask create-sample-services" first.')
            return

        # Create sample service requests with various statuses
        sample_requests_data = [
            {
                'service_name': 'Individual Tax Return',
                'status': 'completed',
                'days_ago': 45,
                'assigned': True,
                'invoice_raised': True,
                'invoice_paid': True,
                'invoice_amount': 350.00,
                'internal_notes': 'FY2024 tax return completed. Client had good documentation.',
                'queries': [
                    {'from': 'accountant', 'message': 'Hi James, I need clarification on the work from home expenses. Can you provide the calculation method you used?', 'days_ago': 40},
                    {'from': 'user', 'message': 'Hi, I used the fixed rate method - 67 cents per hour. I worked from home approximately 1200 hours during the year. I have a log if needed.', 'days_ago': 39},
                    {'from': 'accountant', 'message': 'Thanks for clarifying. That works out to $804. I\'ll include this in your return.', 'days_ago': 38},
                ]
            },
            {
                'service_name': 'Business Activity Statement (BAS)',
                'status': 'processing',
                'days_ago': 10,
                'assigned': True,
                'invoice_raised': True,
                'invoice_paid': True,
                'invoice_amount': 150.00,
                'internal_notes': 'Q2 BAS - All documents received. GST calculations in progress.',
                'queries': []
            },
            {
                'service_name': 'Investment Rental Property',
                'status': 'query_raised',
                'days_ago': 5,
                'assigned': True,
                'invoice_raised': False,
                'invoice_paid': False,
                'internal_notes': 'Waiting for depreciation schedule from client.',
                'queries': [
                    {'from': 'accountant', 'message': 'Hi James, for your rental property at 15 Investment Ave, I noticed you haven\'t provided the depreciation schedule. Do you have one from a quantity surveyor?', 'days_ago': 2},
                ]
            },
            {
                'service_name': 'Company Incorporation',
                'status': 'pending',
                'days_ago': 1,
                'assigned': False,
                'invoice_raised': False,
                'invoice_paid': False,
                'internal_notes': '',
                'queries': []
            },
        ]

        for req_data in sample_requests_data:
            # Find the service
            service = Service.query.filter_by(name=req_data['service_name']).first()
            if not service:
                print(f"Service '{req_data['service_name']}' not found, skipping...")
                continue

            # Check if request already exists for this user and service
            existing = ServiceRequest.query.filter_by(
                user_id=demo_user.id,
                service_id=service.id
            ).first()

            if existing:
                print(f"Request for '{req_data['service_name']}' already exists, skipping...")
                continue

            # Create the service request
            request = ServiceRequest(
                user_id=demo_user.id,
                service_id=service.id,
                status=req_data['status'],
                assigned_accountant_id=accountant.id if req_data['assigned'] and accountant else None,
                invoice_raised=req_data['invoice_raised'],
                invoice_paid=req_data['invoice_paid'],
                invoice_amount=req_data.get('invoice_amount'),
                payment_link='https://payment.example.com/pay/' + str(random.randint(10000, 99999)) if req_data['invoice_raised'] else None,
                internal_notes=req_data['internal_notes'],
                created_at=datetime.utcnow() - timedelta(days=req_data['days_ago'])
            )

            if req_data['status'] == 'completed':
                request.completed_at = datetime.utcnow() - timedelta(days=req_data['days_ago'] - 5)

            db.session.add(request)
            db.session.flush()

            # Add queries/messages
            for q_data in req_data.get('queries', []):
                sender = accountant if q_data['from'] == 'accountant' else demo_user
                if sender:
                    query = Query(
                        service_request_id=request.id,
                        sender_id=sender.id,
                        message=q_data['message'],
                        created_at=datetime.utcnow() - timedelta(days=q_data['days_ago'])
                    )
                    db.session.add(query)

            print(f"Created request: {req_data['service_name']} - Status: {req_data['status']}")

        db.session.commit()
        print('\n--- Demo Data Summary ---')
        print('Demo User: demo@example.com / Demo@123')
        print('Service requests created with various statuses')
        print('Query/message threads added for realistic demo')

    @app.cli.command('seed-clients')
    def seed_clients():
        """Seed clients from Excel file (JPATax Pre-tax Consultation Form)"""
        import os
        from datetime import datetime
        from app.modules.user.models import User, Role
        from app.modules.company.models import Company

        # Try to import openpyxl
        try:
            import openpyxl
        except ImportError:
            print("Error: openpyxl not installed. Run: pip install openpyxl")
            return

        # Look for Excel file in common locations
        excel_paths = [
            '/app/import_data.xlsx',
            '/app/clients.xlsx',
            'import_data.xlsx',
            'clients.xlsx',
        ]

        excel_path = None
        for path in excel_paths:
            if os.path.exists(path):
                excel_path = path
                break

        if not excel_path:
            print("Error: No Excel file found. Place file as 'import_data.xlsx' or 'clients.xlsx'")
            print("Expected locations:", excel_paths)
            return

        print(f"Found Excel file: {excel_path}")

        # Helper functions
        def clean_phone(phone):
            if phone is None:
                return None
            phone = str(phone).strip()
            if phone.endswith('.0'):
                phone = phone[:-2]
            if len(phone) == 9 and phone.isdigit():
                phone = '0' + phone
            return phone

        def clean_tfn(tfn):
            if tfn is None:
                return None
            tfn = str(tfn).strip()
            if tfn.endswith('.0'):
                tfn = tfn[:-2]
            tfn = tfn.replace(' ', '')
            return tfn

        def clean_account_number(acc):
            if acc is None:
                return None
            acc = str(acc).strip()
            if acc.endswith('.0'):
                acc = acc[:-2]
            return acc

        def map_visa_status(visa):
            if visa is None:
                return None
            visa_lower = visa.lower()
            if 'citizen' in visa_lower:
                return 'citizen'
            elif 'permanent' in visa_lower:
                return 'permanent_resident'
            elif 'student' in visa_lower:
                return 'student'
            elif 'working holiday' in visa_lower or '417' in visa_lower or '462' in visa_lower:
                return 'working_holiday'
            elif 'temporary' in visa_lower:
                return 'temporary_resident'
            return 'other'

        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # Get headers
        headers = {}
        for col in range(1, sheet.max_column + 1):
            headers[sheet.cell(row=1, column=col).value] = col

        # Get user role
        user_role = Role.query.filter_by(name=Role.USER).first()
        if not user_role:
            print("Error: User role not found. Run 'flask init-db' first.")
            return

        # Get default company (optional)
        company = Company.query.first()

        imported = 0
        skipped = 0

        print(f"Total rows: {sheet.max_row - 1}")
        print("-" * 50)

        for row in range(2, sheet.max_row + 1):
            try:
                # Get email
                email = sheet.cell(row=row, column=headers.get('Email ID')).value
                if not email:
                    email = sheet.cell(row=row, column=headers.get('Email Address')).value

                if not email:
                    skipped += 1
                    continue

                email = email.strip().lower()

                # Check if user exists
                existing = User.query.filter_by(email=email).first()
                if existing:
                    print(f"Skip (exists): {email}")
                    skipped += 1
                    continue

                # Get data
                first_name = sheet.cell(row=row, column=headers.get('First Name')).value
                last_name = sheet.cell(row=row, column=headers.get('Last / Family Name')).value
                phone = clean_phone(sheet.cell(row=row, column=headers.get('Mobile Number')).value)
                address = sheet.cell(row=row, column=headers.get('Current Residential Address')).value
                visa_status = map_visa_status(sheet.cell(row=row, column=headers.get('Select your current visa status')).value)
                tfn = clean_tfn(sheet.cell(row=row, column=headers.get('Tax File Number (TFN)')).value)
                dob = sheet.cell(row=row, column=headers.get('Date of Birth')).value
                occupation = sheet.cell(row=row, column=headers.get('Occupation')).value
                bsb = sheet.cell(row=row, column=headers.get('BSB')).value
                account_number = clean_account_number(sheet.cell(row=row, column=headers.get('Account Number')).value)
                account_holder = sheet.cell(row=row, column=headers.get('Account Holder Name')).value
                passport_url = sheet.cell(row=row, column=headers.get('Passport Copy (Both pages)')).value
                bank_statement_url = sheet.cell(row=row, column=headers.get('Bank Statement')).value
                driving_licence_url = sheet.cell(row=row, column=headers.get('Australian Driving Licence (both sides)')).value

                # Convert date of birth
                if isinstance(dob, datetime):
                    dob = dob.date()
                else:
                    dob = None

                # Create user
                user = User(
                    email=email,
                    role_id=user_role.id,
                    company_id=company.id if company else None,
                    first_name=first_name.strip() if first_name else None,
                    last_name=last_name.strip() if last_name else None,
                    phone=phone,
                    address=address.strip() if address else None,
                    visa_status=visa_status,
                    tfn=tfn,
                    date_of_birth=dob,
                    occupation=occupation.strip() if occupation else None,
                    bsb=str(bsb).strip() if bsb else None,
                    bank_account_number=account_number,
                    bank_account_holder_name=account_holder.strip() if account_holder else None,
                    passport_url=passport_url,
                    bank_statement_url=bank_statement_url,
                    driving_licence_url=driving_licence_url,
                    terms_accepted=True,
                    terms_accepted_at=datetime.utcnow(),
                    is_active=True,
                    is_verified=False,
                    is_first_login=True
                )
                user.set_password('TempPass123!')

                db.session.add(user)
                imported += 1
                print(f"Import: {email} ({first_name} {last_name})")

            except Exception as e:
                print(f"Error row {row}: {str(e)}")

        db.session.commit()

        print("-" * 50)
        print(f"DONE: Imported {imported}, Skipped {skipped}")
        print("Default password for all imported users: TempPass123!")

    return app
