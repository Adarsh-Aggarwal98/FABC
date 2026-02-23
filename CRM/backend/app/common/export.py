"""
Export utilities for generating CSV and Excel files
"""
import io
import csv
from typing import List, Dict, Any
from flask import Response


def generate_csv(data: List[Dict[str, Any]], columns: List[Dict[str, str]], filename: str) -> Response:
    """
    Generate a CSV file from data.

    Args:
        data: List of dictionaries containing the data
        columns: List of dicts with 'key' and 'title' for column mapping
                 e.g., [{'key': 'email', 'title': 'Email Address'}, ...]
        filename: Name for the downloaded file (without extension)

    Returns:
        Flask Response with CSV content
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header row
    headers = [col['title'] for col in columns]
    writer.writerow(headers)

    # Write data rows
    for row in data:
        row_data = []
        for col in columns:
            value = _get_nested_value(row, col['key'])
            row_data.append(value if value is not None else '')
        writer.writerow(row_data)

    output.seek(0)

    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename={filename}.csv',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )

    return response


def generate_excel(data: List[Dict[str, Any]], columns: List[Dict[str, str]], filename: str) -> Response:
    """
    Generate an Excel file from data.
    Falls back to CSV if openpyxl is not available.

    Args:
        data: List of dictionaries containing the data
        columns: List of dicts with 'key' and 'title' for column mapping
        filename: Name for the downloaded file (without extension)

    Returns:
        Flask Response with Excel content
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'

        # Write headers
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['title'])
            cell.font = cell.font.copy(bold=True)

        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                value = _get_nested_value(row, col['key'])
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for col_idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(ws.cell(row=row, column=col_idx).value or ''))
                for row in range(1, len(data) + 2)
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}.xlsx',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

        return response

    except ImportError:
        # Fall back to CSV if openpyxl not installed
        return generate_csv(data, columns, filename)


def _get_nested_value(obj: Dict, key: str) -> Any:
    """
    Get a value from a nested dictionary using dot notation.

    Args:
        obj: Dictionary to search
        key: Key in dot notation (e.g., 'user.email' or 'address.city')

    Returns:
        The value if found, None otherwise
    """
    keys = key.split('.')
    value = obj

    for k in keys:
        if isinstance(value, dict):
            value = value.get(k)
        else:
            return None

        if value is None:
            return None

    return value


# Column definitions for common exports

USER_EXPORT_COLUMNS = [
    {'key': 'email', 'title': 'Email'},
    {'key': 'first_name', 'title': 'First Name'},
    {'key': 'last_name', 'title': 'Last Name'},
    {'key': 'phone', 'title': 'Phone'},
    {'key': 'address', 'title': 'Address'},
    {'key': 'company_name', 'title': 'Company Name'},
    {'key': 'role', 'title': 'Role'},
    {'key': 'is_active', 'title': 'Active'},
    {'key': 'created_at', 'title': 'Created At'}
]

SERVICE_REQUEST_EXPORT_COLUMNS = [
    {'key': 'id', 'title': 'Request ID'},
    {'key': 'user.email', 'title': 'Client Email'},
    {'key': 'user.full_name', 'title': 'Client Name'},
    {'key': 'service.name', 'title': 'Service'},
    {'key': 'service.category', 'title': 'Category'},
    {'key': 'status', 'title': 'Status'},
    {'key': 'assigned_accountant.full_name', 'title': 'Assigned To'},
    {'key': 'invoice_raised', 'title': 'Invoice Raised'},
    {'key': 'invoice_paid', 'title': 'Invoice Paid'},
    {'key': 'invoice_amount', 'title': 'Invoice Amount'},
    {'key': 'created_at', 'title': 'Created At'},
    {'key': 'completed_at', 'title': 'Completed At'}
]
